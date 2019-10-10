Dir='D:\\Annotations\\'
BCsFileName='BCs.xlsm'
GeomFileName='file.cgns'
ShortAnnotation=True
HorizAxis=0 #0-X, 1-Y, 2-Z
VertAxis=2 #Second axis of the annotation plane: 0-X, 1-Y, 2-Z
FontSize = 55
AnnDist=0.004*FontSize #vertical distance between annotations
RowNum=3 #number of rows for the nearest annotations
TxtColors=[[1.0, 0.0, 0.0],[1.0,1.0,0.0],[0.0, 1.0, 0.0],[0.0,1.0,1.0],[0.0,0.0,1.0]]
#--Settings--------------------------------
import sys
sys.path.append('C:\\Users\\alexey.makrushin\\AppData\\Local\\Continuum\\anaconda2\\Lib\\site-packages')	#for Excel reading
#------------------------------------------
#---------key function for sorting---------
DataTable=[]
BlockNums=[]
def HorizAxisCoord(Num):
	return DataTable[Num]['Coord'][HorizAxis]
#------------------------------------------
Annotations={}
BlocksList=[]
BlockName=''
from paraview.simple import *
#----------------------------------
# reading Excel file
from openpyxl import load_workbook
wb=load_workbook(Dir+BCsFileName)
ws=wb.get_sheet_by_name('Data')
i=2
BlockName=str(ws.cell(i,1).value)
while BlockName!='None':
	if ShortAnnotation:
		Annotations[BlockName]='Zone: '+BlockName+'\n'+\
			'Temp: '+str(ws.cell(i,2).value)+' C\n'+\
			'HTC: ' +str(ws.cell(i,4).value)+' W/m2K\n'+\
			'Press: '+str(ws.cell(i,6).value)+' bar'
	else:
		Annotations[BlockName]='Zone: '+BlockName+'\n'+\
			'Temp: '+str(ws.cell(i,2).value)+' C scaled by '+	str(ws.cell(i,3).value)+'\n'+\
			'HTC: ' +str(ws.cell(i,4).value)+' W/m2K scaled by '+ str(ws.cell(i,5).value)+ '\n'+\
			'Press: '+str(ws.cell(i,6).value)+' bar scaled by '+ str(ws.cell(i,7).value)
	BlocksList.append('/Families/'+BlockName)
	i+=1
	BlockName=str(ws.cell(i,1).value)
wb.close()
#----------------------------------
# create a new 'CGNS Series Reader'
filecgns = CGNSSeriesReader(FileNames=[Dir+GeomFileName])
filecgns.Blocks = BlocksList
# get active view
renderView1 = GetActiveViewOrCreate('RenderView')
# show geometry in view
filecgnsDisplay = Show(filecgns, renderView1)
ColorBy(filecgnsDisplay, ('FIELD', 'vtkBlockColors'))
vtkBlockColorsLUT = GetColorTransferFunction('vtkBlockColors')
#----------------------------------
# data
full_data = servermanager.Fetch(filecgns)				#to get full access to object data
BlockArr=full_data.GetBlock(0).GetBlock(0).GetBlock(1)	#inputs[0] - vtkMultiBlockDataSet
BlocksNum=0
for i in range(0,BlockArr.GetNumberOfBlocks()):
	vtkBlockColorsLUT.IndexedColors[3*(i+1)]=0.75			#color by default
	vtkBlockColorsLUT.IndexedColors[3*(i+1)+1]=0.75			#color by default
	vtkBlockColorsLUT.IndexedColors[3*(i+1)+2]=0.75			#color by default
	SelBlock=BlockArr.GetBlock(i) 						#the last blocks - vtkUnstructuredGrid
#-----Name---------
	Info=BlockArr.GetMetaData(i)
	BlockName=Info.Get(BlockArr.NAME())
#-----Cells--------
	BlockCells=SelBlock.GetCells()
	CellsNum=BlockCells.GetNumberOfCells()
	CellDat=BlockCells.GetData()
# Search of Horizontal Coordinate for an annotation
	Cellj=0
	j=0
	while Cellj<CellsNum:
		CellsNodes=CellDat.GetValue(j)
		for k in range(0,CellsNodes):
			NodeNum=CellDat.GetValue(j+k+1)
			Coord=SelBlock.GetPoints().GetPoint(NodeNum)
			if (j==0 and k==0)or MinHorCoord>Coord[HorizAxis]: MinHorCoord=Coord[HorizAxis]
			if (j==0 and k==0)or MaxHorCoord<Coord[HorizAxis]: MaxHorCoord=Coord[HorizAxis]
		Cellj+=1
		j+=CellsNodes+1
	MdlHorCoord=(MinHorCoord+MaxHorCoord)/2
# Search of rest coordinates for an annotation
	Cellj=0
	j=0
	Flag=True
	while Cellj<CellsNum:
		CellsNodes=CellDat.GetValue(j)
		CellCoord=[0.0,0.0,0.0]
		for k in range(0,CellsNodes):
			NodeNum=CellDat.GetValue(j+k+1)
			Coord=SelBlock.GetPoints().GetPoint(NodeNum)
			CellCoord[0]+=Coord[0]/CellsNodes
			CellCoord[1]+=Coord[1]/CellsNodes
			CellCoord[2]+=Coord[2]/CellsNodes			
			if (k==0)or MinHorCoord>Coord[HorizAxis]: MinHorCoord=Coord[HorizAxis]
			if (k==0)or MaxHorCoord<Coord[HorizAxis]: MaxHorCoord=Coord[HorizAxis]
		if (MinHorCoord-MdlHorCoord)*(MaxHorCoord-MdlHorCoord)<=0:
			if Flag:
				AnnCoord=[CellCoord[0],CellCoord[1],CellCoord[2]]
				Flag=False
			elif AnnCoord[VertAxis]<CellCoord[VertAxis]:
				AnnCoord=[CellCoord[0],CellCoord[1],CellCoord[2]]
		Cellj+=1
		j+=CellsNodes+1
#Table
	if BlockName in Annotations:
		DataTable.append({'Coord':[AnnCoord[0],AnnCoord[1],AnnCoord[2]],'Name':BlockName,'Num':i,'ArrowLength':AnnDist/2})
		BlockNums.append(BlocksNum)
		BlocksNum+=1
vtkBlockColorsLUT.IndexedColors[3*BlockArr.GetNumberOfBlocks()]=0.75			#color by default
vtkBlockColorsLUT.IndexedColors[3*BlockArr.GetNumberOfBlocks()+1]=0.75			#color by default
vtkBlockColorsLUT.IndexedColors[3*BlockArr.GetNumberOfBlocks()+2]=0.75			#color by default
BlockNums.sort(key=HorizAxisCoord)
#----------------------------------
i=0
# show annotations
for j in range(0,BlocksNum):
	Num=BlockNums[j]
	j0=j-RowNum
	if j0<0:j0=0
	for k in BlockNums[j0:j]:
		if abs((DataTable[Num]['Coord'][VertAxis]+DataTable[Num]['ArrowLength'])-(DataTable[k]['Coord'][VertAxis]+DataTable[k]['ArrowLength']))<AnnDist:
			DataTable[Num]['ArrowLength']=DataTable[k]['Coord'][VertAxis]+DataTable[k]['ArrowLength']-DataTable[Num]['Coord'][VertAxis]+AnnDist
# create a new 'Text'
	txt = Text()
	txt.Text = Annotations[DataTable[Num]['Name']]
# show data in view
	txtDisplay=Show(txt, renderView1)
	txtDisplay.FontFamily = 'Courier'
	txtDisplay.FontSize = FontSize
	txtDisplay.Color = TxtColors[i]
	txtDisplay.Opacity = 0.9
	txtDisplay.Bold = 1
	txtDisplay.TextPropMode = 'Flagpole Actor'
	txtDisplay.BasePosition = [DataTable[Num]['Coord'][0],DataTable[Num]['Coord'][1],DataTable[Num]['Coord'][2]]
	txtDisplay.TopPosition = [DataTable[Num]['Coord'][0],DataTable[Num]['Coord'][1],DataTable[Num]['Coord'][2]]
	txtDisplay.TopPosition[VertAxis]+=DataTable[Num]['ArrowLength']
	vtkBlockColorsLUT.IndexedColors[3*(DataTable[Num]['Num']+1)]=TxtColors[i][0]/1.5
	vtkBlockColorsLUT.IndexedColors[3*(DataTable[Num]['Num']+1)+1]=TxtColors[i][1]/1.5
	vtkBlockColorsLUT.IndexedColors[3*(DataTable[Num]['Num']+1)+2]=TxtColors[i][2]/1.5
	i+=1
	if i==len(TxtColors):i=0
#----------------------------------
renderView1.Update()
renderView1.CameraParallelProjection = 1
